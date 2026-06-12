#!/usr/bin/env python3
"""
PI_412 — 3단계: 템플릿(PI_412-프로그램목록.xlsx)을 복사하여 엑셀 생성.

입력:
  - tmp/programs.json
  - 인자 1: 고객사명 (파일명 예약 문자는 _ 치환)
  - 20-deliverables/10-templates/04 구현(PI)/PI_412-프로그램목록.xlsx

출력: 20-deliverables/30-output/04 구현(PI)/PI_412_프로그램목록_{고객사명}.xlsx

전략:
  1. 템플릿을 그대로 복사하여 표지/개정이력 시트 보존.
  2. 프로그램목록_BE 시트: 3행부터 기존 데이터 모두 삭제 → 3행부터 새 데이터 채움.
  3. 프로그램목록_FE 시트: 동일.
  4. auto_filter 범위를 새 데이터 범위로 갱신.
  5. 행 스타일은 템플릿 데이터 첫 행(3행)의 셀 스타일을 복제해서 적용.
"""
from __future__ import annotations

import json
import re
import shutil
import subprocess
import sys
from copy import copy
from pathlib import Path
from typing import Dict, List

REPO_BASE = Path(subprocess.check_output(["git", "rev-parse", "--show-toplevel"], text=True).strip())
OUT_DIR = REPO_BASE / "20-deliverables" / "30-output" / "04 구현(PI)"
TMP_DIR = OUT_DIR / "tmp"
PROGRAMS_JSON = TMP_DIR / "programs.json"
TEMPLATE_XLSX = REPO_BASE / "20-deliverables" / "10-templates" / "04 구현(PI)" / "PI_412-프로그램목록.xlsx"

# 템플릿 컬럼 매핑 (1-base 컬럼 인덱스)
BE_COLS = {
    "lv1": 1, "lv2": 2, "lv3": 3, "lv4": 4, "lv5": 5, "lv6": 6, "lv7": 7,
    "program_id": 8, "program_name": 9,
    "module_name": 10, "module_desc": 11,
    "dev_type": 12, "req_id": 13, "remark": 14,
}
FE_COLS = {
    "lv1": 1, "lv2": 2, "lv3": 3, "lv4": 4, "lv5": 5, "lv6": 6,
    "program_id": 7, "program_name": 8,
    "module_name": 9, "module_desc": 10,
    "dev_type": 11, "req_id": 12, "remark": 13,
}

DATA_START_ROW = 3
HEADER_ROW = 2


def safe_filename(name: str) -> str:
    return re.sub(r'[<>:"|?*\\/]', "_", name).strip() or "고객사미지정"


def clear_sheet_data(ws, last_col: int) -> None:
    """3행부터 max_row까지 모든 셀 값을 지운다 (스타일은 보존)."""
    if ws.max_row < DATA_START_ROW:
        return
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row,
                            min_col=1, max_col=last_col):
        for cell in row:
            cell.value = None


def capture_row_styles(ws, row_idx: int, last_col: int) -> List[dict]:
    """지정 행 셀들의 스타일을 캡처한다."""
    styles = []
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        styles.append({
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
        })
    return styles


def apply_row_styles(ws, row_idx: int, styles: List[dict]) -> None:
    for col_idx, style in enumerate(styles, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.border = style["border"]
        cell.alignment = style["alignment"]
        cell.number_format = style["number_format"]


def fill_sheet(ws, programs: List[Dict], col_map: Dict[str, int]) -> int:
    """시트에 데이터를 채우고 마지막 행 번호를 반환."""
    last_col = max(col_map.values())

    # 데이터 첫 행 스타일 캡처 (없으면 헤더 다음 행을 기본값으로)
    style_row = DATA_START_ROW if ws.max_row >= DATA_START_ROW else HEADER_ROW
    styles = capture_row_styles(ws, style_row, last_col)

    # 기존 데이터 정리
    clear_sheet_data(ws, last_col)

    last_row = HEADER_ROW
    for i, prog in enumerate(programs):
        row_idx = DATA_START_ROW + i
        for key, col in col_map.items():
            val = prog.get(key, "")
            ws.cell(row=row_idx, column=col, value=val if val != "" else None)
        apply_row_styles(ws, row_idx, styles)
        last_row = row_idx

    # auto_filter 범위 갱신 (0건이면 헤더 한 줄만)
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(last_col)
    filter_last_row = last_row if last_row >= DATA_START_ROW else HEADER_ROW
    ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col_letter}{filter_last_row}"

    return last_row


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: 03_generate_excel.py <고객사명>", file=sys.stderr)
        return 2
    customer = safe_filename(sys.argv[1])

    if not PROGRAMS_JSON.exists():
        print(f"[PI_412] programs.json 이 없습니다: {PROGRAMS_JSON}", file=sys.stderr)
        return 2
    if not TEMPLATE_XLSX.exists():
        print(f"[PI_412] 템플릿 파일이 없습니다: {TEMPLATE_XLSX}", file=sys.stderr)
        return 2

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[PI_412] openpyxl 이 설치되어 있지 않습니다", file=sys.stderr)
        return 2

    data = json.loads(PROGRAMS_JSON.read_text(encoding="utf-8"))
    programs = data.get("programs", [])
    be_progs = [p for p in programs if p["section"] == "BE"]
    fe_progs = [p for p in programs if p["section"] == "FE"]
    print(f"[PI_412] 3단계 Excel 생성 시작 (BE {len(be_progs)} / FE {len(fe_progs)})")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"PI_412_프로그램목록_{customer}.xlsx"

    # 1) 템플릿을 그대로 복사
    shutil.copy(TEMPLATE_XLSX, out_path)

    # 2) 워크북 로드 → 두 데이터 시트만 갱신
    wb = load_workbook(out_path)

    if "프로그램목록_BE" in wb.sheetnames:
        ws = wb["프로그램목록_BE"]
        last = fill_sheet(ws, be_progs, BE_COLS)
        print(f"[PI_412]   - 프로그램목록_BE: {len(be_progs)}건 (마지막 행 {last})")
    else:
        print("[PI_412] 경고: 템플릿에 '프로그램목록_BE' 시트가 없습니다", file=sys.stderr)

    if "프로그램목록_FE" in wb.sheetnames:
        ws = wb["프로그램목록_FE"]
        last = fill_sheet(ws, fe_progs, FE_COLS)
        print(f"[PI_412]   - 프로그램목록_FE: {len(fe_progs)}건 (마지막 행 {last})")
    else:
        print("[PI_412] 경고: 템플릿에 '프로그램목록_FE' 시트가 없습니다", file=sys.stderr)

    wb.save(out_path)
    print(f"[PI_412] 완료 → {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
