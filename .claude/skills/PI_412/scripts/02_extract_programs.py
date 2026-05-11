#!/usr/bin/env python3
"""
PI_412 — 2단계: 후보 파일에서 프로그램 메타데이터 추출.

입력: tmp/scan.json
출력: tmp/programs.json

추출 단위: **파일 1개 = 1행** (PI_412 템플릿 형식 준수).

각 항목 필드:
  - section          : "BE" 또는 "FE"
  - lv1 ~ lv7        : 디렉토리 계층 (build prefix 제거 후)
  - program_id       : 가장 깊은 의미 디렉토리 코드 (예: ivad01)
  - program_name     : 도메인 한글명 (템플릿 사전 또는 코멘트)
  - module_name      : 파일 stem (확장자 제거)
  - module_desc      : "{프로그램명} {모듈타입}" 휴리스틱
  - dev_type         : 확장자 (java, xml, vue, scss, ...)
  - req_id, remark   : 빈값
  - path             : 프로젝트 루트 기준 상대 경로
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

REPO_BASE = Path("/mnt/c/zinide/workspace/cloud-wms-doc")
OUT_DIR = REPO_BASE / "output" / "04 구현(PI)"
TMP_DIR = OUT_DIR / "tmp"
SCAN_JSON = TMP_DIR / "scan.json"
OUT_JSON = TMP_DIR / "programs.json"
TEMPLATE_XLSX = REPO_BASE / "template" / "04 구현(PI)" / "PI_412-프로그램목록.xlsx"

KOREAN = re.compile(r"[가-힣]")

# Build/소스 prefix — Lv 분해에서 제거
SRC_PREFIXES_BE = [
    ["src", "main", "java"],
    ["src", "main", "kotlin"],
    ["src", "main", "resources"],
    ["src", "main", "resource"],
    ["src", "main"],
    ["src"],
]
SRC_PREFIXES_FE = [
    ["src"],
]


def read_text(path: Path, max_bytes: int = 200_000) -> str:
    try:
        with open(path, "rb") as f:
            return f.read(max_bytes).decode("utf-8", errors="ignore")
    except Exception:
        return ""


def korean_first(*candidates: str) -> str:
    for c in candidates:
        if c and KOREAN.search(c):
            return c
    return ""


# ---------- 디렉토리 → Lv 분해 ----------

def strip_prefix(parts: List[str], prefixes: List[List[str]]) -> List[str]:
    """parts에서 가장 긴 일치 prefix를 제거한다."""
    for pref in sorted(prefixes, key=len, reverse=True):
        n = len(pref)
        for i in range(0, max(0, len(parts) - n) + 1):
            if parts[i:i + n] == pref:
                return parts[i + n:]
    return parts


def split_lv(rel: str, section: str, max_lv: int) -> Tuple[List[str], str]:
    """
    상대경로를 Lv 배열과 파일명으로 분해한다.
    BE는 src/main/java 같은 build prefix 제거 후의 디렉토리(예: be/iv3000/ivad01).
    FE는 src 제거 전 그대로(예: src/api, src/components/...) — 템플릿 관례.
    """
    parts = rel.split("/")
    file_name = parts[-1]
    dir_parts = parts[:-1]

    if section == "BE":
        # src/main/java 등 build prefix 제거 → 첫 항목이 그대로 Lv1 (be / fw / bm 등)
        lvs = strip_prefix(dir_parts, SRC_PREFIXES_BE)
    else:
        # FE는 src 자체를 Lv1로 둔다 (템플릿이 'src'를 Lv1로 사용)
        lvs = list(dir_parts)

    lvs = lvs[:max_lv] + [""] * max(0, max_lv - len(lvs))
    return lvs, file_name


# ---------- 파일 유형 추정 (모듈설명용) ----------

# Java/XML: 파일명에서 도메인 코드 prefix를 제거한 suffix를 모듈타입으로 사용
MODULE_TYPE_KO = {
    "controller":  "Controller",
    "comp":        "Comp",
    "comput":      "CompUtil",
    "computil":    "CompUtil",
    "txcomp":      "TxComp",
    "dao":         "Dao",
    "mapper":      "Mapper",
    "sqlmapper":   "SqlMapper",
    "service":     "Service",
    "request":     "Request",
    "response":    "Response",
    "vo":          "VO",
    "dto":         "DTO",
    "config":      "Config",
    "util":        "Util",
    "exception":   "Exception",
    "filter":      "Filter",
    "interceptor": "Interceptor",
    "handler":     "Handler",
    "excelcontroller": "Excel Controller",
    "excelcomp":   "Excel Comp",
}


def split_module_type(stem: str, dev_type: str) -> str:
    """
    IVAD01Controller → 'Controller'
    IVAD01CompUtil   → 'CompUtil'
    PushAlarmMapper  → 'Mapper'
    매칭이 없으면 stem의 카멜 분해 마지막 토큰을 반환.
    """
    s = stem
    # 우선순위: 알려진 suffix 매칭 (긴 것부터)
    lower = s.lower()
    for key in sorted(MODULE_TYPE_KO.keys(), key=len, reverse=True):
        if lower.endswith(key):
            return MODULE_TYPE_KO[key]
    # 카멜 분해
    tokens = re.findall(r"[A-Z][a-z0-9]*|[a-z0-9]+", s)
    if tokens:
        return tokens[-1]
    return dev_type


# ---------- 템플릿 사전 (도메인코드 → 한글명) ----------

def load_template_dictionary() -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    PI_412 템플릿에서 사전을 추출한다.
      - domain_dict: {도메인코드 → 한글 프로그램명}  (예: 'ivad01' → '재고조정')
      - module_desc_dict: {모듈명 → 한글 설명} (예: 'IVAD01Controller' → '재고조정 Controller')
    템플릿이 없으면 빈 dict 반환.
    """
    domain_dict: Dict[str, str] = {}
    module_desc_dict: Dict[str, str] = {}
    if not TEMPLATE_XLSX.exists():
        return domain_dict, module_desc_dict
    try:
        from openpyxl import load_workbook
    except ImportError:
        return domain_dict, module_desc_dict

    try:
        wb = load_workbook(TEMPLATE_XLSX, data_only=True, read_only=True)
    except Exception:
        return domain_dict, module_desc_dict

    # 도메인 매핑은 두 단계로 채택한다:
    #  1) Controller 행이 있으면 그 program_name 사용 (도메인을 가장 잘 대표함)
    #  2) 없으면 첫 번째 한글 program_name fallback
    primary: Dict[str, str] = {}    # pid → name (Controller 우선)
    fallback: Dict[str, str] = {}   # pid → name (그 외)

    def absorb(pid: str, pname: str, mname: str, mdesc: str) -> None:
        if pid and pname and KOREAN.search(pname):
            key = pid.lower()
            if mname.endswith("Controller") or mname.endswith("Service"):
                primary[key] = pname
            else:
                fallback.setdefault(key, pname)
        if mname and mdesc:
            module_desc_dict.setdefault(mname, mdesc)

    # BE 시트: H=프로그램ID, I=프로그램명, J=모듈명, K=모듈설명
    if "프로그램목록_BE" in wb.sheetnames:
        ws = wb["프로그램목록_BE"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or row[7] is None:
                continue
            pid = str(row[7]).strip()
            pname = str(row[8]).strip() if row[8] else ""
            mname = str(row[9]).strip() if len(row) > 9 and row[9] else ""
            mdesc = str(row[10]).strip() if len(row) > 10 and row[10] else ""
            absorb(pid, pname, mname, mdesc)

    # FE 시트: G=프로그램ID, H=프로그램명, I=모듈명, J=모듈설명
    if "프로그램목록_FE" in wb.sheetnames:
        ws = wb["프로그램목록_FE"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or row[6] is None:
                continue
            pid = str(row[6]).strip()
            pname = str(row[7]).strip() if row[7] else ""
            mname = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            mdesc = str(row[9]).strip() if len(row) > 9 and row[9] else ""
            absorb(pid, pname, mname, mdesc)

    # primary 우선 적용, 없는 키만 fallback에서 보충
    domain_dict.update(fallback)
    domain_dict.update(primary)

    return domain_dict, module_desc_dict


# ---------- 코멘트에서 한글 발췌 ----------

def korean_from_text(text: str) -> str:
    """파일 상단 200줄 이내의 한글 코멘트 첫 줄을 발췌."""
    if not text:
        return ""
    head = text[:3000]
    # /** ... */ 블록
    m = re.search(r"/\*+([\s\S]+?)\*+/", head)
    if m:
        for ln in m.group(1).splitlines():
            ln = re.sub(r"^\s*\*+\s?", "", ln).strip()
            if ln and KOREAN.search(ln) and not ln.startswith("@"):
                return ln
    # // 또는 # 한글 코멘트
    for ln in head.splitlines():
        s = ln.strip()
        if s.startswith("//") and KOREAN.search(s):
            return s.lstrip("/").strip()
        if s.startswith("#") and KOREAN.search(s):
            return s.lstrip("#").strip()
    # vue/html 주석
    m = re.search(r"<!--([\s\S]+?)-->", head)
    if m:
        for ln in m.group(1).splitlines():
            s = ln.strip()
            if s and KOREAN.search(s):
                return s
    return ""


# ---------- 메인 추출 ----------

def extract_program_id_be(lvs: List[str], stem: str) -> str:
    """
    BE 프로그램ID는 가장 깊은 의미 디렉토리 코드.
    템플릿 예: be/comm → 'comm', be/iv3000/ivad01 → 'ivad01', be/iv3000/ivad01/bean → 'ivad01' (한 단계 위)
    bean/excel/test 같은 보조 디렉토리는 건너뛴다.
    """
    skip = {"bean", "excel", "test", "tests", "util", "utils", "vo", "dto", "exception", "config", ""}
    # lvs[0]은 'be'. lvs[1:]에서 뒤에서부터 의미 있는 코드 탐색
    for v in reversed(lvs[1:]):
        if v and v.lower() not in skip:
            return v
    # fallback: lvs[1] 또는 stem
    return lvs[1] if len(lvs) > 1 and lvs[1] else stem


def extract_program_id_fe(lvs: List[str], stem: str) -> str:
    """FE 프로그램ID는 파일 stem 그대로 (템플릿 관례)."""
    return stem


def build_program_record_be(rel: str, abs_path: Path,
                            domain_dict: Dict[str, str],
                            module_desc_dict: Dict[str, str]) -> Dict:
    lvs, fname = split_lv(rel, "BE", 7)
    stem = Path(fname).stem
    dev_type = Path(fname).suffix.lstrip(".").lower()
    program_id = extract_program_id_be(lvs, stem)

    # 프로그램명: 사전 → 코멘트 → 빈값
    program_name = domain_dict.get(program_id.lower(), "")
    if not program_name:
        program_name = korean_first(korean_from_text(read_text(abs_path)))

    module_name = stem
    # 모듈설명: 사전 우선 → 휴리스틱
    module_desc = module_desc_dict.get(module_name, "")
    if not module_desc:
        mtype = split_module_type(stem, dev_type)
        if program_name:
            module_desc = f"{program_name} {mtype}".strip()
        else:
            module_desc = mtype

    return {
        "section": "BE",
        "lv1": lvs[0], "lv2": lvs[1], "lv3": lvs[2], "lv4": lvs[3],
        "lv5": lvs[4], "lv6": lvs[5], "lv7": lvs[6],
        "program_id": program_id,
        "program_name": program_name,
        "module_name": module_name,
        "module_desc": module_desc,
        "dev_type": dev_type,
        "req_id": "",
        "remark": "",
        "path": rel,
    }


def build_program_record_fe(rel: str, abs_path: Path,
                            domain_dict: Dict[str, str],
                            module_desc_dict: Dict[str, str]) -> Dict:
    lvs, fname = split_lv(rel, "FE", 6)
    stem = Path(fname).stem
    dev_type = Path(fname).suffix.lstrip(".").lower()
    program_id = extract_program_id_fe(lvs, stem)

    program_name = domain_dict.get(program_id.lower(), "")
    if not program_name:
        program_name = korean_first(korean_from_text(read_text(abs_path)))

    module_name = stem
    module_desc = module_desc_dict.get(module_name, "")
    if not module_desc:
        # FE는 파일명을 그대로 설명에 쓰는 사례가 많음
        module_desc = stem

    return {
        "section": "FE",
        "lv1": lvs[0], "lv2": lvs[1], "lv3": lvs[2], "lv4": lvs[3],
        "lv5": lvs[4], "lv6": lvs[5], "lv7": "",
        "program_id": program_id,
        "program_name": program_name,
        "module_name": module_name,
        "module_desc": module_desc,
        "dev_type": dev_type,
        "req_id": "",
        "remark": "",
        "path": rel,
    }


def main() -> int:
    if not SCAN_JSON.exists():
        print(f"[PI_412] scan.json 이 없습니다. 1단계를 먼저 실행하세요: {SCAN_JSON}", file=sys.stderr)
        return 2
    scan = json.loads(SCAN_JSON.read_text(encoding="utf-8"))
    target = Path(scan["target_dir"])
    files: Dict[str, List[str]] = scan["files"]

    domain_dict, module_desc_dict = load_template_dictionary()
    print(f"[PI_412] 템플릿 사전 로드: 도메인 {len(domain_dict)}개 / 모듈설명 {len(module_desc_dict)}개")

    programs: List[Dict] = []
    print("[PI_412] 2단계 추출 시작 (파일 단위)")

    spring_files = files.get("spring", []) or []
    fe_files = files.get("frontend", []) or []
    py_files = files.get("python", []) or []

    # 1차 패스: Controller 파일에서 도메인 코드 → 한글명 보강 (템플릿 사전 미커버 도메인용)
    code_domain_hint: Dict[str, str] = {}
    for rel in spring_files:
        fname = Path(rel).name
        if not fname.endswith("Controller.java") and not fname.endswith("Controller.kt"):
            continue
        text = read_text(target / rel)
        kr = korean_first(korean_from_text(text))
        if not kr:
            continue
        # program_id 산출 (build_program_record_be와 동일 로직)
        lvs, _ = split_lv(rel, "BE", 7)
        stem = Path(fname).stem
        pid = extract_program_id_be(lvs, stem).lower()
        if pid and pid not in domain_dict:
            code_domain_hint.setdefault(pid, kr)
    if code_domain_hint:
        print(f"[PI_412] 코드 코멘트 도메인 보강: {len(code_domain_hint)}개")
        for k, v in code_domain_hint.items():
            domain_dict.setdefault(k, v)

    # 2차 패스: 파일별 레코드 빌드
    for rel in spring_files:
        abs_path = target / rel
        rec = build_program_record_be(rel, abs_path, domain_dict, module_desc_dict)
        programs.append(rec)
    for rel in py_files:
        abs_path = target / rel
        rec = build_program_record_be(rel, abs_path, domain_dict, module_desc_dict)
        programs.append(rec)
    for rel in fe_files:
        abs_path = target / rel
        rec = build_program_record_fe(rel, abs_path, domain_dict, module_desc_dict)
        programs.append(rec)

    # 정렬: BE → FE; lv1~lv7 → module_name
    section_order = {"BE": 0, "FE": 1}
    programs.sort(key=lambda x: (
        section_order.get(x["section"], 9),
        x["lv1"], x["lv2"], x["lv3"], x["lv4"], x["lv5"], x["lv6"], x.get("lv7", ""),
        x["module_name"], x["dev_type"],
    ))

    be = sum(1 for p in programs if p["section"] == "BE")
    fe = sum(1 for p in programs if p["section"] == "FE")
    print(f"[PI_412] 합계: BE {be}건 / FE {fe}건 / 총 {len(programs)}건")

    out = {
        "extracted_at": datetime.now().isoformat(timespec="seconds"),
        "target_dir": str(target),
        "programs": programs,
        "stats": {"be": be, "fe": fe, "total": len(programs)},
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[PI_412] 완료 → {OUT_JSON}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
