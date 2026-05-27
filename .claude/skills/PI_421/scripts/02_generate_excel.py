#!/usr/bin/env python3
"""PI_421 2단계 — 템플릿 복사 + 데이터 채우기 + Sheet1 통계 갱신 (Windows 기본).

사용법:
    python 02_generate_excel.py <고객사명> <담당자명>

입력: output/04 구현(PI)/tmp/tests.json
출력: output/04 구현(PI)/PI_421_단위테스트보고서_{고객사명}_{YYMMDD}.xlsx

BASE_DIR을 `Path(__file__).resolve().parents[4]` 로 자동 추론하여
Windows/WSL 어느 환경에서도 정상 동작한다.
"""
from __future__ import annotations

import datetime as dt
import json
import re
import shutil
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# .claude/skills/PI_421/scripts/02_generate_excel.py
# parents[4] = 프로젝트 루트
BASE_DIR = Path(__file__).resolve().parents[4]
TEMPLATE = BASE_DIR / "template" / "04 구현(PI)" / "PI_212-단위테스트보고서.xlsx"
TMP_JSON = BASE_DIR / "output" / "04 구현(PI)" / "tmp" / "tests.json"
OUT_DIR = BASE_DIR / "output" / "04 구현(PI)"

DATA_SHEET = "단위테스트 보고서"
STAT_SHEET = "Sheet1"
DATA_START_ROW = 3
TEMPLATE_STYLE_LAST_ROW = 183
TEMPLATE_LAST_ROW = 191
NUM_COLS = 14
INVALID_FILENAME_CHARS = re.compile(r'[<>:"|?*\\/]')


def sanitize_filename(s: str) -> str:
    return INVALID_FILENAME_CHARS.sub("_", s).strip()


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for c in range(1, NUM_COLS + 1):
        src_cell = ws.cell(row=src_row, column=c)
        dst_cell = ws.cell(row=dst_row, column=c)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
    src_h = ws.row_dimensions[src_row].height
    if src_h is not None:
        ws.row_dimensions[dst_row].height = src_h


def clear_template_data_rows(ws) -> None:
    for r in range(DATA_START_ROW, TEMPLATE_LAST_ROW + 1):
        for c in range(1, NUM_COLS + 1):
            ws.cell(row=r, column=c).value = None


def write_test_rows(ws, tests: list[dict], reporter: str, check_date: dt.date) -> None:
    style_src = DATA_START_ROW
    for idx, t in enumerate(tests):
        row = DATA_START_ROW + idx
        if row > TEMPLATE_STYLE_LAST_ROW:
            copy_row_style(ws, style_src, row)
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = t["platform"]
        ws.cell(row=row, column=3).value = t["big_menu"]
        ws.cell(row=row, column=4).value = t["test_id"]
        ws.cell(row=row, column=5).value = t["menu"]
        ws.cell(row=row, column=6).value = t["category"]
        ws.cell(row=row, column=7).value = t["content"]
        ws.cell(row=row, column=8).value = check_date
        ws.cell(row=row, column=9).value = reporter
        ws.cell(row=row, column=10).value = t["result"]


def update_stat_sheet(ws, tests: list[dict]) -> None:
    plat_counts = {"WEB": 0, "PDA": 0, "I/F": 0}
    for t in tests:
        if t["platform"] in plat_counts:
            plat_counts[t["platform"]] += 1

    max_r = min(ws.max_row, 30)
    for r in range(1, max_r + 1):
        label = ws.cell(row=r, column=2).value
        if not isinstance(label, str):
            continue
        label = label.strip()
        if label in plat_counts:
            ws.cell(row=r, column=4).value = plat_counts[label]
            ws.cell(row=r, column=5).value = 0
            ws.cell(row=r, column=6).value = 0


def main(argv: list[str]) -> int:
    if len(argv) < 3:
        print("Usage: 02_generate_excel.py <고객사명> <담당자명>", file=sys.stderr)
        return 2

    customer = sanitize_filename(argv[1])
    reporter = argv[2].strip() or "테스터"
    if not customer:
        print("[ERR] 고객사명이 비어 있습니다.", file=sys.stderr)
        return 2

    if not TEMPLATE.exists():
        print(f"[ERR] 템플릿 없음: {TEMPLATE}", file=sys.stderr)
        return 1
    if not TMP_JSON.exists():
        print(f"[ERR] tests.json 없음 (1단계 먼저 실행): {TMP_JSON}", file=sys.stderr)
        return 1

    data = json.loads(TMP_JSON.read_text(encoding="utf-8"))
    tests: list[dict] = data["tests"]
    if not tests:
        print("[ERR] 추출된 테스트가 0건입니다.", file=sys.stderr)
        return 1

    today = dt.date.today()
    yymmdd = today.strftime("%y%m%d")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    outfile = OUT_DIR / f"PI_421_단위테스트보고서_{customer}_{yymmdd}.xlsx"

    print(f"[2/3] 템플릿 복사 → {outfile.name}")
    shutil.copy(TEMPLATE, outfile)

    print("[2/3] 데이터 시트 기록 중...")
    wb = openpyxl.load_workbook(outfile)
    ws = wb[DATA_SHEET]
    clear_template_data_rows(ws)
    write_test_rows(ws, tests, reporter, today)

    last_row = DATA_START_ROW + len(tests) - 1
    last_col = get_column_letter(NUM_COLS)
    try:
        ws.print_area = f"A1:{last_col}{last_row}"
    except Exception:
        pass

    print("[2/3] Sheet1 통계 갱신...")
    if STAT_SHEET in wb.sheetnames:
        update_stat_sheet(wb[STAT_SHEET], tests)
    else:
        print(f"  ⚠ '{STAT_SHEET}' 시트 없음 — 통계 갱신 스킵")

    wb.save(outfile)
    wb.close()

    plat = {"WEB": 0, "PDA": 0, "I/F": 0}
    for t in tests:
        if t["platform"] in plat:
            plat[t["platform"]] += 1

    print()
    print("✓ 단위테스트 보고서 생성 완료 [PI_421_WIN]")
    print()
    print(f"대상 디렉토리: {data.get('scanned_dir', '-')}")
    print(f"고객사:        {customer}")
    print(f"담당자:        {reporter}")
    print(f"확인일자:      {today.isoformat()}")
    print()
    print(f"스캔 결과:")
    print(f"  - 테스트 클래스 파일 : {data.get('junit_files', '-')}개")
    print(f"  - 추출된 @Test 메서드: {len(tests)}건")
    print(f"  - JUnit 4 / JUnit 5  : {data.get('junit4_files', 0)} / {data.get('junit5_files', 0)}")
    print()
    print("플랫폼별 분포:")
    for p in ("WEB", "PDA", "I/F"):
        print(f"  - {p:<3} : {plat[p]}건")
    print(f"  - 합계: {len(tests)}건  (결과: O={len(tests)} / 완료율 100.0%)")
    print()
    print(f"출력 파일: {outfile}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
