#!/usr/bin/env python3
"""PI_421 1단계 — JUnit 테스트 메서드 스캔 (Windows 기본).

사용법:
    python 01_scan_tests.py <백엔드_디렉토리>

출력:
    deliverables/30-output/04 구현(PI)/tmp/tests.json

BASE_DIR을 `Path(__file__).resolve().parents[4]` 로 자동 추론하여
Windows/WSL 어느 환경에서도 정상 동작한다.
"""
from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

# .claude/skills/PI_421/scripts/01_scan_tests.py
# parents[0]=scripts  parents[1]=PI_421  parents[2]=skills
# parents[3]=.claude  parents[4]=<프로젝트 루트>
BASE_DIR = Path(__file__).resolve().parents[4]
TMP_DIR = BASE_DIR / "deliverables" / "30-output" / "04 구현(PI)" / "tmp"
OUT_JSON = TMP_DIR / "tests.json"

EXCLUDE_DIRS = {
    "node_modules", "dist", "build", "target", ".git",
    ".gradle", ".mvn", "bin", "obj", "out",
    "__pycache__", ".venv", "venv", ".idea", ".vscode",
}

# lv2(또는 lv1) 키워드 → (대메뉴, 플랫폼)
LV2_MAP = [
    # 기준정보
    ("md8000", "기준정보", "WEB"),
    ("master", "기준정보", "WEB"),
    # 입고
    ("iw1000", "입고", "WEB"),
    ("inbound", "입고", "WEB"),
    ("receive", "입고", "WEB"),
    # 반품
    ("rt2000", "반품", "WEB"),
    ("return", "반품", "WEB"),
    # 재고
    ("iv3000", "재고", "WEB"),
    ("iv3100", "재고", "WEB"),
    ("iv3200", "재고", "WEB"),
    ("inventory", "재고", "WEB"),
    ("stock", "재고", "WEB"),
    # 출고
    ("ow5000", "출고", "WEB"),
    ("outbound", "출고", "WEB"),
    ("delivery", "출고", "WEB"),
    ("shipping", "출고", "WEB"),
    # 시스템관리
    ("mm9200", "시스템관리", "WEB"),
    ("sm9000", "시스템관리", "WEB"),
    ("ss9300", "시스템관리", "WEB"),
    ("admin", "시스템관리", "WEB"),
    ("system", "시스템관리", "WEB"),
    # 인터페이스
    ("if9100", "인터페이스", "I/F"),
    ("interface", "인터페이스", "I/F"),
]

INTERFACE_HINTS = ("if9100", "sif/", "/sif/", "interface")

DOMAIN_CODE_RE = re.compile(r"^(?:ZTEST_)?([A-Z]{2,5}[0-9]{2,3}M?)")

MODULE_SUFFIX_KO = {
    "Comp": "Comp",
    "Controller": "Controller",
    "Dao": "Dao",
    "Mapper": "Mapper",
    "Service": "Service",
    "SUITE": "Suite",
    "Util": "Util",
}

METHOD_PATTERNS = [
    (re.compile(r"^findAll[_A-Z]?"), "전체 목록 조회"),
    (re.compile(r"^findById[_A-Z]?"), "단건 조회"),
    (re.compile(r"^getById[_A-Z]?"), "단건 조회"),
    (re.compile(r"^find[_A-Z]"), "조회"),
    (re.compile(r"^get[_A-Z]"), "조회"),
    (re.compile(r"^search[_A-Z]?"), "검색"),
    (re.compile(r"^select[_A-Z]?"), "조회"),
    (re.compile(r"^create[_A-Z]?"), "등록"),
    (re.compile(r"^register[_A-Z]?"), "등록"),
    (re.compile(r"^save[_A-Z]?"), "저장"),
    (re.compile(r"^insert[_A-Z]?"), "등록"),
    (re.compile(r"^update[_A-Z]?"), "수정"),
    (re.compile(r"^modify[_A-Z]?"), "수정"),
    (re.compile(r"^delete[_A-Z]?"), "삭제"),
    (re.compile(r"^remove[_A-Z]?"), "삭제"),
    (re.compile(r"^test_search"), "검색"),
    (re.compile(r"^test_select"), "조회"),
    (re.compile(r"^test_insert"), "등록"),
    (re.compile(r"^test_update"), "수정"),
    (re.compile(r"^test_delete"), "삭제"),
    (re.compile(r"^test_"), "기능 테스트"),
]


def normalize_path(p: str) -> str:
    """입력 경로를 현재 OS에 맞게 정규화.

    Windows에서는 `/mnt/<drive>/...` 형태의 WSL 경로를 `<DRIVE>:\\...` 로 변환한다.
    Linux/WSL에서는 `C:\\...` 형태를 `/mnt/c/...` 로 변환한다.
    그 외는 입력 그대로 사용.
    """
    p = p.strip().strip('"').strip("'")
    # WSL 경로 → Windows 경로 (Windows에서 실행될 때)
    m = re.match(r"^/mnt/([a-zA-Z])/(.*)$", p)
    if m and os.name == "nt":
        drive = m.group(1).upper()
        rest = m.group(2).replace("/", "\\")
        return f"{drive}:\\{rest}"
    # Windows 경로 → WSL 경로 (Linux에서 실행될 때)
    if re.match(r"^[A-Za-z]:[\\/]", p) and os.name != "nt":
        drive = p[0].lower()
        rest = p[2:].replace("\\", "/").lstrip("/")
        return f"/mnt/{drive}/{rest}"
    return p


def find_test_files(root: Path) -> list[Path]:
    """JUnit 테스트 후보 파일 수집."""
    candidates: list[Path] = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in EXCLUDE_DIRS]
        for fn in filenames:
            if not (fn.endswith(".java") or fn.endswith(".kt")):
                continue
            rel_str = str(Path(dirpath)).replace("\\", "/")
            in_test_root = "/src/test/java" in rel_str or "/src/test/kotlin" in rel_str
            name_hits_test = (
                fn.endswith("Test.java") or fn.endswith("Tests.java")
                or fn.startswith("Test") or fn.startswith("ZTEST_")
                or fn.endswith("Test.kt") or fn.endswith("Tests.kt")
            )
            if in_test_root or name_hits_test:
                candidates.append(Path(dirpath) / fn)
    return candidates


_BLOCK_COMMENT_RE = re.compile(r"/\*.*?\*/", re.DOTALL)
_TEST_ANN_RE = re.compile(
    r"@(?:org\.junit\.jupiter\.api\.|org\.junit\.)?Test\b"
    r"(?:\s*\([^)]*\))?",
    re.MULTILINE,
)
_DISPLAY_RE = re.compile(r'@DisplayName\s*\(\s*"([^"]*)"\s*\)')
_METHOD_RE = re.compile(
    r"\b(?:void|[\w<>\[\]\.]+)\s+([A-Za-z_]\w*)\s*\([^)]*\)\s*(?:throws[^{;]*)?\s*\{",
)
_KEYWORDS = {"if", "for", "while", "switch", "return", "new", "catch", "synchronized", "try"}

BUILTIN_NAME_DICT = {
    "MDBZ01": "사업장관리",
    "MDCT01": "거래처관리",
    "MDPD01": "품목관리",
    "MDWH01": "창고관리",
    "MDUS01": "사용자관리",
    "MDLC01": "위치관리",
    "MDCR01": "운반사관리",
    "MDSP01": "공급처관리",
    "MDLP01": "로케이션프린트",
    "IWRQ01": "입고예정",
    "IWPC01": "입고처리",
    "IWLB01": "입고라벨",
    "RTRQ01": "반품예정",
    "RTPC01": "반품처리",
    "IVMV01": "재고이동",
    "IVAD01": "재고조정",
    "IVIO01": "수불현황",
    "IVPD01": "재고조회-품목별",
    "IVWH01": "재고조회-창고별",
    "IVMC01": "재고현황",
    "IVST01": "재고현황",
    "IVSK01": "SKU재고",
    "SKHT01": "SKU이력",
    "OBRQ01": "출고예정",
    "OBPC01": "출고처리",
    "OWRC01": "출고처리",
    "OWRB01": "출고예정",
    "OWPC01": "출고처리",
    "DLPC01": "송장처리",
    "DLCX01": "배송취소",
    "DLPB01": "배송점검",
    "LDPC01": "상차확정",
    "SMCC01": "공통코드",
    "SMMN01": "메뉴관리",
    "SMMG01": "시스템관리",
    "SMBD01": "게시판",
    "USCD01": "사용자코드",
    "STRG01": "스캔규칙",
    "STSC01": "스캔체크",
    "SCRG01": "스캔규칙",
    "ALSH01": "알람경보",
    "ALST01": "알람이력",
    "PDST01": "품목상태",
    "BRSC01": "바코드스캔",
    "LGAP01": "로그인",
    "LGCO01": "로그아웃",
    "LGER01": "로그인오류",
    "LGMN01": "로그메뉴",
    "IFBH01": "IF배치이력",
    "SHA256": "SHA256암호화",
}


def strip_block_comments(src: str) -> str:
    return _BLOCK_COMMENT_RE.sub(lambda m: " " * len(m.group(0)), src)


def extract_tests_from_file(path: Path, project_root: Path) -> tuple[list[dict], int]:
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        print(f"[WARN] read failed: {path} → {e}", file=sys.stderr)
        return [], 0

    junit_v = 0
    if "org.junit.jupiter.api.Test" in raw:
        junit_v = 5
    elif re.search(r"import\s+org\.junit\.Test\b", raw):
        junit_v = 4

    src = strip_block_comments(raw)

    lines = src.split("\n")
    masked: list[str] = []
    for ln in lines:
        stripped = ln.lstrip()
        if stripped.startswith("//"):
            masked.append(" " * len(ln))
        else:
            idx = ln.find("//")
            if idx >= 0:
                masked.append(ln[:idx] + " " * (len(ln) - idx))
            else:
                masked.append(ln)
    src = "\n".join(masked)

    out: list[dict] = []
    test_positions = [m.start() for m in _TEST_ANN_RE.finditer(src)]
    if not test_positions:
        return [], junit_v

    rel_file = path.relative_to(project_root)
    class_name = path.stem
    package_path = _derive_package_path(rel_file)

    for pos in test_positions:
        tail = src[pos:pos + 600]
        ctx = src[max(0, pos - 200):pos + 600]
        d = _DISPLAY_RE.search(ctx)
        display_name = d.group(1).strip() if d else None
        method_name = None
        for m in _METHOD_RE.finditer(tail):
            cand = m.group(1)
            if cand in _KEYWORDS:
                continue
            method_name = cand
            break
        if not method_name:
            continue
        out.append({
            "class_file": str(rel_file).replace("\\", "/"),
            "class_name": class_name,
            "method_name": method_name,
            "display_name": display_name,
            "package_path": package_path,
            "junit_version": junit_v or 5,
        })
    return out, junit_v


def _derive_package_path(rel_file: Path) -> str:
    parts = rel_file.parts
    if "java" in parts:
        i = parts.index("java")
        sub = parts[i + 1 : -1]
        return "/".join(sub)
    if "kotlin" in parts:
        i = parts.index("kotlin")
        sub = parts[i + 1 : -1]
        return "/".join(sub)
    return "/".join(rel_file.parts[:-1])


def classify_menu(package_path: str, class_file: str) -> tuple[str, str]:
    pkg = package_path.lower()
    cf = class_file.lower()
    for h in INTERFACE_HINTS:
        if h in pkg or h in cf:
            return ("인터페이스", "I/F")
    if pkg.startswith("fw/") or pkg.startswith("test/") or pkg == "test":
        return ("공통", "WEB")
    if pkg.startswith("bm/") or "/bm/" in cf:
        platform = "PDA"
    else:
        platform = "WEB"

    segments = pkg.split("/")
    candidates = segments[1:3] if len(segments) >= 2 else segments
    norm_candidates = []
    for c in candidates:
        norm_candidates.append(c)
        if c.endswith("m") and len(c) > 3:
            norm_candidates.append(c[:-1])

    for cand in norm_candidates:
        for key, big, _plat in LV2_MAP:
            if cand == key or cand.startswith(key):
                return (big, platform)

    if segments and segments[0] in {"be", "bm"} and len(segments) > 1:
        return (segments[1] or "기타", platform)
    return ("기타", platform)


def extract_domain_code(class_name: str) -> str | None:
    m = DOMAIN_CODE_RE.match(class_name)
    if m:
        return m.group(1)
    return None


def humanize_method(method_name: str) -> str:
    for rx, label in METHOD_PATTERNS:
        if rx.match(method_name):
            return f"{label} - {method_name}"
    return method_name


def make_content(test: dict) -> str:
    if test.get("display_name"):
        return test["display_name"]
    return humanize_method(test["method_name"])


def load_program_name_dict() -> dict[str, str]:
    tpl = BASE_DIR / "deliverables" / "10-templates" / "04 구현(PI)" / "PI_412-프로그램목록.xlsx"
    result: dict[str, str] = {}
    if not tpl.exists():
        return result
    try:
        import openpyxl
        wb = openpyxl.load_workbook(tpl, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if "프로그램목록" not in sheet_name:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not row or len(row) < 9:
                    continue
                pid = row[7]
                pnm = row[8]
                if not (isinstance(pid, str) and isinstance(pnm, str)):
                    continue
                pid = pid.strip().upper()
                pnm = pnm.strip()
                if not pid or not pnm or pnm in {"공통"}:
                    continue
                result.setdefault(pid, pnm)
        wb.close()
    except Exception as e:
        print(f"[INFO] PI_412 사전 학습 스킵: {e}", file=sys.stderr)
    return result


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Usage: 01_scan_tests.py <백엔드_디렉토리>", file=sys.stderr)
        return 2

    root = Path(normalize_path(argv[1])).resolve()
    if not root.exists() or not root.is_dir():
        print(f"[ERR] 디렉토리가 존재하지 않습니다: {root}", file=sys.stderr)
        return 2

    TMP_DIR.mkdir(parents=True, exist_ok=True)

    print(f"[1/3] 스캔 시작: {root}")
    files = find_test_files(root)
    print(f"  • JUnit 후보 파일: {len(files)}건")
    if not files:
        print("[ERR] Java/Kotlin 테스트 파일이 발견되지 않았습니다.", file=sys.stderr)
        return 1

    name_dict = load_program_name_dict()
    if name_dict:
        print(f"  • PI_412 한글명 사전: {len(name_dict)}개 학습")

    all_tests: list[dict] = []
    junit4 = junit5 = 0
    for i, f in enumerate(files, 1):
        if i % 50 == 0 or i == len(files):
            print(f"  ... {i}/{len(files)} 파일 처리", flush=True)
        tests, jv = extract_tests_from_file(f, root)
        if jv == 4:
            junit4 += 1
        elif jv == 5:
            junit5 += 1
        for t in tests:
            big_menu, platform = classify_menu(t["package_path"], t["class_file"])
            domain_code = extract_domain_code(t["class_name"])
            if domain_code:
                base_code = domain_code.rstrip("M") if domain_code.endswith("M") else domain_code
                ko = (
                    name_dict.get(domain_code)
                    or name_dict.get(base_code)
                    or BUILTIN_NAME_DICT.get(domain_code)
                    or BUILTIN_NAME_DICT.get(base_code)
                )
                if not ko:
                    bare = re.sub(r"^ZTEST_", "", t["class_name"])
                    bare = re.sub(r"(Comp|Controller|Dao|Mapper|Service|Util|Suite|Biz)$", "", bare)
                    ko = bare
                menu = f"{ko}({domain_code})"
            else:
                menu = re.sub(r"^ZTEST_", "", t["class_name"])
            t.update({
                "big_menu": big_menu,
                "platform": platform,
                "menu": menu,
                "category": "기능",
                "content": make_content(t),
                "result": "O",
            })
            all_tests.append(t)

    PLAT_ORDER = {"WEB": 0, "PDA": 1, "I/F": 2}
    all_tests.sort(key=lambda t: (
        PLAT_ORDER.get(t["platform"], 9),
        t["big_menu"],
        t["class_name"],
    ))

    for idx, t in enumerate(all_tests, start=1):
        t["test_id"] = f"WMS-BE-{idx:03d}"

    result = {
        "scanned_dir": str(root),
        "junit_files": len(files),
        "test_count": len(all_tests),
        "junit4_files": junit4,
        "junit5_files": junit5,
        "tests": all_tests,
    }

    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  • 추출된 @Test 메서드: {len(all_tests)}건")
    print(f"  • JUnit 4 / JUnit 5 파일: {junit4} / {junit5}")
    for p in ("WEB", "PDA", "I/F"):
        n = sum(1 for t in all_tests if t["platform"] == p)
        print(f"  • {p:<3}: {n}건")
    print(f"[OK] tests.json → {OUT_JSON}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
