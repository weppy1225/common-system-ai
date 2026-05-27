"""
에이전트 3: output/02 분析(RA)/tmp/requirements.json 을 읽어
요구사항정의서 xlsx 파일을 생성한다.

사용법: python3 scripts/03_generate_excel.py
작업 디렉토리: git rev-parse --show-toplevel (동적 감지)
"""
import openpyxl, json, shutil, os, glob, math, subprocess
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import Counter

BASE  = subprocess.check_output(["git", "rev-parse", "--show-toplevel"], text=True).strip()
TODAY = datetime.now().strftime("%y%m%d")

# ── 경로 탐색 (한/중 혼용 인코딩 대응) ──────────────────────────────
def find_dir(pattern):
    results = glob.glob(pattern)
    return results[0] if results else None

outdir = find_dir(os.path.join(BASE, "output", "*RA*"))
if not outdir:
    outdir = os.path.join(BASE, "output", "02 분析(RA)")
    os.makedirs(outdir, exist_ok=True)

tmp = os.path.join(outdir, "tmp")

# ── requirements.json 로드 ───────────────────────────────────────────
with open(os.path.join(tmp, "requirements.json"), encoding="utf-8") as f:
    data = json.load(f)

company = data["company"]
reqs    = data["requirements"]
out_path = os.path.join(outdir, f"RA.222-요구사항정의서_{company}_{TODAY}.xlsx")

# ── 템플릿 복사 ──────────────────────────────────────────────────────
tpl_path = find_dir(os.path.join(BASE, "template", "*RA*", "RA.314*.xlsx"))
if tpl_path and os.path.exists(tpl_path):
    shutil.copy2(tpl_path, out_path)
    wb = openpyxl.load_workbook(out_path)
else:
    wb = openpyxl.Workbook()
    wb.active.title = "요구사항정의서"

ws = wb["요구사항정의서"]

# ── 1. 기존 데이터 행 완전 삭제 (값 초기화가 아닌 행 삭제) ─────────
total_rows = ws.max_row
if total_rows >= 4:
    ws.delete_rows(4, total_rows - 3)  # 헤더 3행 보존, 나머지 삭제

# ── 2. 스타일 정의 ───────────────────────────────────────────────────
thin      = Side(style="thin", color="CCCCCC")
border    = Border(left=thin, right=thin, top=thin, bottom=thin)
fill_even = PatternFill("solid", fgColor="F9FAFB")
font_base = Font(name="맑은 고딕", size=9)

CONTENT_COL_WIDTH = 58   # 내용 컬럼 실제 사용 너비(units)
LINE_HEIGHT_PT    = 14   # font 9 기준 줄당 높이(pt)
ROW_PADDING_PT    = 6    # 셀 상하 여백(pt)

def calc_row_height(text, col_units=CONTENT_COL_WIDTH):
    """wrap_text 기준으로 필요한 행 높이(pt)를 계산한다."""
    if not text:
        return LINE_HEIGHT_PT + ROW_PADDING_PT
    lines = str(text).split("\n")
    total = 0
    for line in lines:
        if not line.strip():
            total += 1
            continue
        # 한글 2unit, ASCII/숫자/공백 1unit
        w = sum(2 if ord(c) > 127 else 1 for c in line)
        total += max(1, math.ceil(w / col_units))
    return max(LINE_HEIGHT_PT + ROW_PADDING_PT, total * LINE_HEIGHT_PT + ROW_PADDING_PT)

# ── 3. 데이터 기입 ──────────────────────────────────────────────────
for i, req in enumerate(reqs):
    row_idx = i + 4
    cells = [
        (1,  i + 1),
        (2,  req["id"]),
        (3,  req["name"]),
        (4,  req["section_name"]),
        (5,  req["content"]),
        (6,  req["requester"]),
        (7,  req["date"]),
        (8,  req["priority"]),
        (9,  req["acceptance"]),
        (10, req["remark"]),
    ]
    fill = fill_even if i % 2 == 1 else None
    for col, val in cells:
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border = border
        cell.font   = font_base
        cell.alignment = Alignment(
            horizontal="center" if col in (1, 6, 7, 8, 9) else "left",
            vertical="top",
            wrap_text=(col == 5),
        )
        if fill:
            cell.fill = fill
    ws.row_dimensions[row_idx].height = calc_row_height(req["content"])

# ── 4. 컬럼 너비 자동 조정 ──────────────────────────────────────────
# 한글 1자 ≈ 2 unit, ASCII 1자 ≈ 1 unit (Excel 문자폭 기준)
def text_width(value):
    if value is None:
        return 0
    # 줄바꿈이 있으면 가장 긴 줄 기준
    lines = str(value).split("\n")
    return max(sum(2 if ord(c) > 127 else 1 for c in line) for line in lines)

# 컬럼별 최소·최대 너비 (unit)
COL_MIN = {1: 5, 2: 14, 3: 18, 4: 9,  5: 40, 6: 12, 7: 13, 8: 8,  9: 14, 10: 20}
COL_MAX = {1: 8, 2: 18, 3: 28, 4: 10, 5: 60, 6: 16, 7: 13, 8: 9,  9: 16, 10: 40}

col_width = {c: COL_MIN[c] for c in COL_MIN}

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        col = cell.column
        if col not in col_width:
            continue
        w = text_width(cell.value) + 2  # 좌우 여백 2 unit
        col_width[col] = min(max(col_width[col], w), COL_MAX[col])

for col, width in col_width.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── 5. 표지 업체명 갱신 ─────────────────────────────────────────────
if "표지" in wb.sheetnames:
    for row in wb["표지"].iter_rows():
        for cell in row:
            if cell.value and "WMS" in str(cell.value) and "구축" in str(cell.value):
                cell.value = f"{company} WMS 구축 프로젝트"
                break

# ── 6. 개정이력 날짜 갱신 ───────────────────────────────────────────
if "개정이력" in wb.sheetnames:
    for row in wb["개정이력"].iter_rows():
        for cell in row:
            if cell.value == "v1.0":
                wb["개정이력"].cell(row=cell.row, column=5).value = datetime.now().strftime("%Y.%m.%d")
                break

# ── 7. 요구사항 일람표 건수 갱신 ────────────────────────────────────
if "요구사항 일람표" in wb.sheetnames:
    ws_list = wb["요구사항 일람표"]
    section_count = Counter(r["section_code"] for r in reqs)
    code_map = {
        "CO":"공통","MD":"기준","IW":"입고","RM":"반품",
        "IV":"재고","OW":"출고","IF":"I/F","PDA":"PDA","ERR":"예외"
    }
    name_to_code = {v: k for k, v in code_map.items()}
    for row in ws_list.iter_rows(min_row=3):
        for cell in row:
            if cell.column == 4 and cell.value in name_to_code:
                code = name_to_code[str(cell.value)]
                ws_list.cell(row=cell.row, column=6).value = section_count.get(code, 0)

# ── 저장 ────────────────────────────────────────────────────────────
wb.save(out_path)
print(f"생성 완료: {out_path}")
print(f"요구사항 총 {len(reqs)}건 입력")

section_order = ["CO","MD","IW","RM","IV","OW","IF","PDA","ERR"]
section_label = {"CO":"공통","MD":"기준","IW":"입고","RM":"반품","IV":"재고",
                 "OW":"출고","IF":"I/F","PDA":"PDA","ERR":"예외"}
section_count = Counter(r["section_code"] for r in reqs)
for code in section_order:
    if code in section_count:
        print(f"  {section_label[code]}({code}): {section_count[code]}건")
