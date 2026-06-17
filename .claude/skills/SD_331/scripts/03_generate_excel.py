#!/usr/bin/env python3
"""
SD_331 3단계 - 추출된 schema를 SD.212-테이블정의서 템플릿에 채워 Excel 생성.

입력:
    deliverables/30-output/03 설계(SD)/tmp/schema.json
    template/03 설계(SD)/SD.212-테이블정의서.xlsx

출력:
    deliverables/30-output/03 설계(SD)/SD.212-테이블정의서_{DB명}_{YYMMDD}.xlsx

전략:
1. 템플릿 복사
2. 첫 번째 테이블 시트(원본의 'MDM_사업장')를 블루프린트로 정리:
   - Table info 값 비움
   - 각 섹션의 데이터 행을 1행만 남기고 정리(머지 제거)
3. 다른 모든 샘플 테이블 시트 삭제
4. Table List 시트 데이터 행 정리
5. 추출된 각 테이블에 대해:
   - 블루프린트 시트를 copy_worksheet으로 복제
   - 시트명을 logical name으로 변경 (Excel 31자 제한 준수)
   - Table info 채움
   - 각 섹션 데이터 채움(필요 시 insert_rows로 행 확장)
6. 블루프린트 시트 삭제
7. Table List 채움
8. 저장
"""

import datetime
import json
import os
import re
import shutil
import sys
from copy import copy
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[4]
TEMPLATE_PATH = BASE_DIR / "template/03 설계(SD)/SD.212-테이블정의서.xlsx"
OUTPUT_DIR = BASE_DIR / "deliverables/30-output/03 설계(SD)"
TMP_DIR = OUTPUT_DIR / "tmp"
SCHEMA_FILE = TMP_DIR / "schema.json"

SECTION_LABELS = ["Table info", "Column info", "Index info", "Constraint info", "FK info", "FK info (PK Side)"]
INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def safe_sheet_name(name, used):
    """Excel 시트명 규칙(31자, 금지 문자 제거, 중복 회피)."""
    if not name:
        name = "Sheet"
    name = INVALID_SHEET_CHARS.sub("_", str(name)).strip()
    if not name:
        name = "Sheet"
    base = name[:31]
    candidate = base
    n = 2
    while candidate.lower() in {x.lower() for x in used}:
        suffix = f"_{n}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(candidate)
    return candidate


def find_label_row(ws, label, max_row=None):
    if max_row is None:
        max_row = ws.max_row
    for r in range(1, max_row + 1):
        v = ws.cell(r, 1).value
        if v == label:
            return r
    return None


def find_section_rows(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v in SECTION_LABELS:
            out[v] = r
    return out


def _safe_set_value(ws, coord, value):
    """머지된 셀(MergedCell)은 main 셀에만 값을 set한다.

    coord가 가리키는 셀이 MergedCell이면, 그 셀이 속한 머지 범위의 좌상단 셀에 값을 set.
    """
    from openpyxl.cell.cell import MergedCell
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        # 어떤 머지에 속해 있는지 찾는다
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                ws.cell(mr.min_row, mr.min_col).value = value
                return
        # 못 찾으면 무시
        return
    cell.value = value


def remove_merges_in_range(ws, min_row, max_row):
    """주어진 행 범위와 겹치는 merged_cells를 모두 제거."""
    to_remove = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= min_row and mr.max_row <= max_row:
            to_remove.append(str(mr))
        elif mr.min_row <= max_row and mr.max_row >= min_row:
            to_remove.append(str(mr))
    for mr in to_remove:
        ws.unmerge_cells(mr)


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def copy_row_style(ws, src_row, dst_row, max_col):
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def cleanup_blueprint(ws):
    """블루프린트 시트 정리: 모든 데이터 영역을 데이터 1행만 남기고 정리.

    정리 후 구조 (데이터 행 모두 1행):
      row N    : 'Table info'
      row N+1~6: Table info 라벨/값 (값은 비움)
      ...
      row M    : 'Column info'
      row M+1  : 컬럼 헤더
      row M+2  : 데이터 1행 (덮어쓸 자리)
      row M+3  : 'Index info'
      row M+4  : 인덱스 헤더
      row M+5  : 데이터 1행
      ... 이하 동일
    """
    # 끝 섹션부터 정리해야 행 인덱스가 어긋나지 않음
    sections_in_order = ["FK info (PK Side)", "FK info", "Constraint info", "Index info", "Column info"]
    label_to_next = {
        "FK info (PK Side)": None,  # 마지막 섹션 - end of sheet
        "FK info": "FK info (PK Side)",
        "Constraint info": "FK info",
        "Index info": "Constraint info",
        "Column info": "Index info",
    }

    for sect in sections_in_order:
        s = find_section_rows(ws)
        if sect not in s:
            continue
        label_row = s[sect]
        header_row = label_row + 1
        data_start = label_row + 2
        next_label = label_to_next.get(sect)
        if next_label and next_label in s:
            data_end = s[next_label] - 1
        else:
            data_end = ws.max_row

        # 데이터 영역에 걸친 머지 제거 후 데이터 행을 1개만 남기고 모두 삭제
        if data_end >= data_start:
            remove_merges_in_range(ws, data_start, data_end)
            # 데이터 1행만 남기고 모두 삭제
            keep_row = data_start
            # 1행을 비워둔다 (데이터는 새로 채워질 자리, 스타일은 유지)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(keep_row, col)
                cell.value = None
            if data_end > data_start:
                ws.delete_rows(data_start + 1, data_end - data_start)

    # Table info 값 비우기 (라벨 셀은 유지). 머지된 셀은 main 셀에만 None 설정.
    s = find_section_rows(ws)
    table_info_row = s.get("Table info", 1)
    column_info_row = s.get("Column info", 12)
    for r in range(table_info_row + 1, column_info_row):
        for col_letter in ("C", "D", "F", "G"):
            _safe_set_value(ws, f"{col_letter}{r}", None)


def fill_table_info(ws, table, db_info, today):
    """Table info 섹션 채우기. 위치는 라벨 텍스트로 동적 탐색."""
    # 라벨이 들어 있는 위치를 찾는다 (B 컬럼)
    s = find_section_rows(ws)
    table_info_row = s.get("Table info", 1)
    column_info_row = s.get("Column info", 12)

    for r in range(table_info_row + 1, column_info_row):
        b_label = ws.cell(r, 2).value
        e_label = ws.cell(r, 5).value
        if b_label == "System Name":
            _safe_set_value(ws, f"C{r}", db_info.get("database", ""))
        elif b_label == "Sub-system Name":
            _safe_set_value(ws, f"C{r}", "")
        elif b_label == "Schema Name":
            _safe_set_value(ws, f"C{r}", table.get("schema", "") or db_info.get("schema", ""))
        elif b_label == "Logical Table Name":
            _safe_set_value(ws, f"C{r}", table.get("logical_name", ""))
        elif b_label == "Pyshical Table Name" or b_label == "Physical Table Name":
            _safe_set_value(ws, f"C{r}", table.get("physical_name", ""))
        elif b_label == "Remark":
            _safe_set_value(ws, f"C{r}", table.get("comment", ""))

        if e_label == "Author":
            _safe_set_value(ws, f"F{r}", os.environ.get("USER") or os.environ.get("USERNAME") or "")
        elif e_label == "Created On":
            _safe_set_value(ws, f"F{r}", today)
        elif e_label == "Modified On":
            _safe_set_value(ws, f"F{r}", today)
        elif e_label == "RDBMS":
            _safe_set_value(ws, f"F{r}", db_info.get("driver", ""))


def fill_section(ws, section_label, items, col_count, mapper):
    """섹션의 데이터를 채운다. 블루프린트는 데이터 1행을 비워두고 있다고 가정."""
    label_row = find_label_row(ws, section_label)
    if label_row is None:
        return
    data_start = label_row + 2

    if not items:
        # 데이터 0개: 1행을 빈 채로 둠
        for col in range(1, col_count + 1):
            ws.cell(data_start, col).value = None
        return

    # 첫 행 채우기
    values = mapper(0, items[0])
    for col_idx, v in enumerate(values, 1):
        ws.cell(data_start, col_idx).value = v

    # 추가 행
    if len(items) > 1:
        ws.insert_rows(data_start + 1, amount=len(items) - 1)
        for offset in range(1, len(items)):
            copy_row_style(ws, data_start, data_start + offset, max_col=col_count)
            values = mapper(offset, items[offset])
            for col_idx, v in enumerate(values, 1):
                ws.cell(data_start + offset, col_idx).value = v


def get_pk_columns(table):
    for idx in table.get("indexes", []):
        if idx.get("is_pk"):
            return [c.strip() for c in (idx.get("columns") or "").split(",") if c.strip()]
    for con in table.get("constraints", []):
        if (con.get("type") or "").upper() in ("PRIMARY KEY", "P"):
            return [c.strip() for c in (con.get("definition") or "").split(",") if c.strip()]
    return []


def fill_table_sheet(ws, table, db_info, today):
    fill_table_info(ws, table, db_info, today)

    pk_cols = set(get_pk_columns(table))

    def col_mapper(i, c):
        nn = c.get("not_null")
        if c["physical_name"] in pk_cols:
            null_str = "Yes (PK)"
        elif nn:
            null_str = "Yes"
        else:
            null_str = ""
        return [
            i + 1,
            c.get("logical_name", ""),
            c.get("physical_name", ""),
            c.get("data_type", ""),
            null_str,
            c.get("default", "") or "",
            c.get("comment", "") or "",
        ]

    fill_section(ws, "Column info", table.get("columns", []), col_count=7, mapper=col_mapper)

    fill_section(ws, "Index info", table.get("indexes", []), col_count=7,
                 mapper=lambda i, x: [i + 1, x.get("name", ""), x.get("columns", ""), None,
                                       "Yes" if x.get("is_pk") else "",
                                       "Yes" if x.get("is_unique") else "",
                                       x.get("remark", "")])

    fill_section(ws, "Constraint info", table.get("constraints", []), col_count=4,
                 mapper=lambda i, c: [i + 1, c.get("name", ""), c.get("type", ""), c.get("definition", "")])

    fill_section(ws, "FK info", table.get("fks", []), col_count=7,
                 mapper=lambda i, f: [i + 1, f.get("name", ""), f.get("columns", ""), None,
                                       f.get("ref_table", ""), None, f.get("ref_columns", "")])

    fill_section(ws, "FK info (PK Side)", table.get("fks_pk_side", []), col_count=7,
                 mapper=lambda i, f: [i + 1, f.get("name", ""), f.get("columns", ""), None,
                                       f.get("ref_table", ""), None, f.get("ref_columns", "")])


def fill_table_list(wb, tables):
    ws = wb["Table List"]
    # 헤더 다음 데이터 행 모두 정리
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    # 헤더 행(2행) 스타일을 유지하기 위해 첫 데이터 행은 없는 셈으로 둠
    # 새 데이터 채움
    for i, t in enumerate(tables, 1):
        r = 2 + i
        ws.cell(r, 1).value = i
        ws.cell(r, 2).value = t.get("logical_name", "")
        ws.cell(r, 3).value = t.get("physical_name", "")
        ws.cell(r, 4).value = t.get("comment", "")


def main():
    if not SCHEMA_FILE.exists():
        print(f"[SD_331] schema 파일 없음: {SCHEMA_FILE}", file=sys.stderr)
        sys.exit(2)
    if not TEMPLATE_PATH.exists():
        print(f"[SD_331] 템플릿 없음: {TEMPLATE_PATH}", file=sys.stderr)
        sys.exit(2)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[SD_331] openpyxl 미설치. 설치: python3 -m pip install --user openpyxl", file=sys.stderr)
        sys.exit(2)

    schema = json.loads(SCHEMA_FILE.read_text(encoding="utf-8"))
    db_info = schema.get("db", {})
    tables = schema.get("tables", [])
    if not tables:
        print("[SD_331] 테이블 0개. 추출된 데이터가 없음.", file=sys.stderr)
        sys.exit(2)

    today = datetime.date.today()
    yymmdd = today.strftime("%y%m%d")
    db_name = db_info.get("database") or "db"
    safe_dbname = re.sub(r"[^\w\-]+", "_", db_name)
    output_path = OUTPUT_DIR / f"SD.212-테이블정의서_{safe_dbname}_{yymmdd}.xlsx"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE_PATH, output_path)
    print(f"[SD_331] 템플릿 복사: {output_path}")

    wb = load_workbook(output_path)

    # 1. 첫 번째 테이블 시트를 블루프린트로 만든다 (Table List를 제외한 첫 시트)
    sheet_names = wb.sheetnames
    table_sheet_names = [n for n in sheet_names if n != "Table List"]
    if not table_sheet_names:
        print("[SD_331] 템플릿에 테이블 시트가 없음.", file=sys.stderr)
        sys.exit(2)

    blueprint_name = table_sheet_names[0]
    print(f"[SD_331] 블루프린트 시트: {blueprint_name}")
    blueprint = wb[blueprint_name]

    # 2. 다른 모든 테이블 시트 삭제 (블루프린트는 마지막에 삭제)
    for name in table_sheet_names[1:]:
        del wb[name]

    # 3. 블루프린트 정리
    cleanup_blueprint(blueprint)
    # 정리된 블루프린트 시트명 변경 - 임시명
    blueprint.title = "__BLUEPRINT__"

    # 4. Table List 정리 (헤더만 남김)
    fill_table_list_headers_only(wb)

    # 5. 추출된 각 테이블에 대해 시트 생성
    used_names = {"Table List", "__BLUEPRINT__"}
    today_str = today.strftime("%Y-%m-%d")
    for t in tables:
        new_ws = wb.copy_worksheet(blueprint)
        # 시트명 정하기 (logical name 우선, 없으면 physical)
        desired = t.get("logical_name") or t.get("physical_name") or "Sheet"
        new_name = safe_sheet_name(desired, used_names)
        new_ws.title = new_name
        fill_table_sheet(new_ws, t, db_info, today_str)

    # 6. 블루프린트 삭제
    del wb["__BLUEPRINT__"]

    # 7. Table List 채움
    fill_table_list(wb, tables)

    # 8. 저장
    wb.save(output_path)

    total_cols = sum(len(t.get("columns", [])) for t in tables)
    total_idx = sum(len(t.get("indexes", [])) for t in tables)
    total_con = sum(len(t.get("constraints", [])) for t in tables)
    total_fk = sum(len(t.get("fks", [])) for t in tables)

    print(f"[SD_331] 생성 완료: {output_path}")
    print(f"  테이블: {len(tables)}개")
    print(f"  컬럼: {total_cols}, 인덱스: {total_idx}, 제약조건: {total_con}, FK: {total_fk}")


def fill_table_list_headers_only(wb):
    """Table List에서 헤더(2행) 이후를 모두 비운다."""
    ws = wb["Table List"]
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)


if __name__ == "__main__":
    main()
