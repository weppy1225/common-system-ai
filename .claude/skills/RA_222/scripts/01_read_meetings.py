"""
에이전트 1: deliverables/20-sources/RA.212/ 폴더의 모든 xlsx 파일을 읽어
deliverables/30-output/02 분석(RA)/tmp/meeting_raw.json 에 저장한다.

사용법: python scripts/01_read_meetings.py
작업 디렉토리: /mnt/c/zinide/workspace/cloud-wms-doc
"""
import openpyxl, json, os, glob, subprocess
from datetime import datetime

base = subprocess.check_output(["git", "rev-parse", "--show-toplevel"], text=True).strip()
input_dir = os.path.join(base, "deliverables", "20-sources", "RA.212")
out_dir = os.path.join(base, "deliverables", "30-output", "02 분석(RA)", "tmp")
os.makedirs(out_dir, exist_ok=True)

result = {"files": []}

for xlsx_path in sorted(glob.glob(os.path.join(input_dir, "*.xlsx"))):
    fname = os.path.basename(xlsx_path)
    if fname.startswith("~$"):
        continue
    doc_type = "회의록" if "회의록" in fname else "메뉴현황" if "메뉴현황" in fname else "기타"
    file_entry = {"filename": fname, "doc_type": doc_type, "sheets": []}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        file_entry["error"] = str(e)
        result["files"].append(file_entry)
        continue

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            cleaned = []
            for cell in row:
                if cell is None:
                    cleaned.append("")
                elif isinstance(cell, datetime):
                    cleaned.append(cell.strftime("%Y-%m-%d"))
                else:
                    cleaned.append(str(cell).strip())
            if any(v for v in cleaned):
                rows_data.append(cleaned)
        file_entry["sheets"].append({"sheet_name": sname, "rows": rows_data})

    result["files"].append(file_entry)

out_path = os.path.join(out_dir, "meeting_raw.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"저장 완료: {out_path}")
print(f"파일 수: {len(result['files'])}")
for fi in result["files"]:
    sheets_info = ", ".join(s["sheet_name"] for s in fi.get("sheets", []))
    print(f"  - {fi['filename']} [{fi['doc_type']}] 시트: {sheets_info}")
