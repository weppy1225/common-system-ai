#!/usr/bin/env python3
"""
[SD_332] 3단계 — Excel 산출물 생성

템플릿: template/04 구현(PI)/PI_113-공통코드정의서.xlsx
- 시트:
    표지, 개정이력, 3.코드그룹, 4.상세코드, 그룹SQL, 상세SQL

채우는 위치:
  3.코드그룹  헤더 7행 / 데이터 8행~
    A=biz_seq  B=comm_h_cd  C=comm_h_nm  D=(빈값)
    G=user_cd_yn  H=inout_cd  I=비고(use_yn='N' → '미사용', 그 외 빈값)

  4.상세코드  헤더 2~3행 / 데이터 4행~
    A=biz_seq(str)  B=comm_h_cd  C=comm_d_cd  D=comm_d_nm
    E=ref_h_cd  F=ref_d_cd  G=disp_no  H=disp_yn  I=use_yn

  그룹SQL    행마다 INSERT 수식. SQL row n → 코드그룹 row n+7
  상세SQL    행마다 INSERT 수식. SQL row n → 상세코드 row n+3

입력: deliverables/30-output/04 구현(PI)/tmp/common_codes.json
출력: deliverables/30-output/04 구현(PI)/PI_113-공통코드정의서_{YYMMDD}.xlsx
"""

from __future__ import annotations

import json
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

BASE = Path(__file__).resolve().parents[4]
TEMPLATE = BASE / "template" / "04 구현(PI)" / "PI_113-공통코드정의서.xlsx"
TMP_DIR = BASE / "output" / "04 구현(PI)" / "tmp"
OUT_DIR = BASE / "output" / "04 구현(PI)"
IN_FILE = TMP_DIR / "common_codes.json"


# ---------------------------------------------------------------------------
# SQL 수식 빌더
# ---------------------------------------------------------------------------

def make_group_formula(n: int) -> str:
    """그룹SQL 시트의 n번째 행 수식 (1-based). 코드그룹 시트의 n+7행을 참조."""
    s = n + 7
    return (
        f"=IF(ISBLANK('3.코드그룹'!H{s}), "
        f"CONCATENATE(B{n}, '3.코드그룹'!A{s}, \",\", \"'\", '3.코드그룹'!B{s}, "
        f"\"' ,'\", '3.코드그룹'!C{s},\"', '\", '3.코드그룹'!G{s}, \"', \", \"NULL\", 그룹SQL!C{n}), "
        f"CONCATENATE(B{n}, '3.코드그룹'!A{s}, \",\", \"'\", '3.코드그룹'!B{s}, "
        f"\"' ,'\", '3.코드그룹'!C{s},\"', '\", '3.코드그룹'!G{s}, \"', '\", "
        f"'3.코드그룹'!H{s}, \"'\", 그룹SQL!C{n}))"
    )


GROUP_B_TEXT = "INSERT INTO SM_COMM_H (biz_seq, comm_h_cd, comm_h_nm, user_cd_yn, inout_cd, use_yn, reg_dt, reg_id) VALUES ("
GROUP_C_TEXT = ",'Y', getDate(), 'administrator' );"


def make_detail_formula(n: int) -> str:
    """상세SQL 시트의 n번째 행 수식 (1-based). 상세코드 시트의 n+3행을 참조."""
    s = n + 3
    return (
        f"=CONCATENATE(B{n}, '4.상세코드'!A{s}, \", '\", '4.상세코드'!B{s}, "
        f"\"', '\", '4.상세코드'!C{s}, \"', '\", '4.상세코드'!D{s}, \"',\", "
        f"IF(ISBLANK('4.상세코드'!E{s}), \"NULL\", CONCATENATE(\"'\", '4.상세코드'!E{s}, \"'\")), \",\", "
        f"IF(ISBLANK('4.상세코드'!F{s}), \"NULL\", CONCATENATE(\"'\", '4.상세코드'!F{s}, \"'\")), \",\", "
        f"'4.상세코드'!G{s}, \", '\", '4.상세코드'!H{s}, \"', '\", '4.상세코드'!I{s}, \"'\", 상세SQL!C{n})"
    )


DETAIL_B_TEXT = (
    "INSERT INTO SM_COMM_D (biz_seq, comm_h_cd, comm_d_cd, comm_d_nm, ref_h_cd, ref_d_cd, "
    "disp_no, disp_yn, use_yn, reg_id, reg_dt) VALUES ("
)
DETAIL_C_TEXT = ", 'administrator', getDate());"


# ---------------------------------------------------------------------------
# Excel 작성
# ---------------------------------------------------------------------------

def clear_data_rows(ws, start_row: int, max_col: int) -> int:
    """start_row 이상의 모든 셀의 value를 None으로 비운다. 반환값은 비워진 마지막 행."""
    last = ws.max_row
    if last < start_row:
        return last
    for r in range(start_row, last + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None
    return last


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    """src_row의 셀 스타일과 행 높이를 dst_row로 복사한다. (값은 보존)"""
    if src_row == dst_row:
        return
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if not src.has_style:
            continue
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format
    src_height = ws.row_dimensions[src_row].height if src_row in ws.row_dimensions else None
    if src_height is not None:
        ws.row_dimensions[dst_row].height = src_height


def trim_trailing_rows(ws, keep_until: int) -> None:
    """keep_until 이후의 모든 행을 물리적으로 삭제한다.

    keep_until 이후 행에 걸친 병합 셀은 먼저 unmerge한 뒤 삭제한다.
    keep_until 행 자체와 그 위쪽의 병합 셀은 그대로 둔다.
    """
    last = ws.max_row
    if last <= keep_until:
        return
    to_unmerge = [rng for rng in list(ws.merged_cells.ranges) if rng.min_row > keep_until]
    for rng in to_unmerge:
        ws.unmerge_cells(str(rng))
    ws.delete_rows(keep_until + 1, last - keep_until)


def fill_groups(ws, groups: list[dict[str, Any]]) -> None:
    """3.코드그룹 시트에 그룹 데이터 작성."""
    HEADER_ROW = 7
    START_ROW = 8
    MAX_COL = 9
    template_last_row = ws.max_row  # 템플릿 데이터 영역의 마지막 행 (스타일 보유 한계)
    clear_data_rows(ws, START_ROW, MAX_COL)

    for idx, g in enumerate(groups):
        r = START_ROW + idx
        biz_seq = g.get("biz_seq")
        ws.cell(row=r, column=1, value=int(biz_seq) if biz_seq is not None else None)  # A
        ws.cell(row=r, column=2, value=g.get("comm_h_cd"))  # B
        ws.cell(row=r, column=3, value=g.get("comm_h_nm"))  # C
        # D=코드설명: DB 미존재 → 빈값
        ws.cell(row=r, column=4, value=None)
        # G=user_cd_yn  H=inout_cd
        ws.cell(row=r, column=7, value=g.get("user_cd_yn"))
        inout_cd = g.get("inout_cd")
        ws.cell(row=r, column=8, value=inout_cd if inout_cd not in (None, "") else None)
        # I=비고 (use_yn='N' → '미사용')
        use_yn = (g.get("use_yn") or "Y").upper()
        ws.cell(row=r, column=9, value="미사용" if use_yn == "N" else None)
        # 템플릿 행 수를 초과한 새 행은 첫 데이터 행 스타일/행 높이를 복사한다.
        if r > template_last_row:
            copy_row_style(ws, START_ROW, r, MAX_COL)

    last_data_row = START_ROW + len(groups) - 1 if groups else START_ROW - 1
    # 데이터 행마다 D:F 병합 (템플릿 시각 유지). 기존 병합은 unmerge 후 재적용.
    existing_dfm = {rng.min_row: rng for rng in list(ws.merged_cells.ranges)
                    if rng.min_col == 4 and rng.max_col == 6}
    for r in range(START_ROW, last_data_row + 1):
        if r in existing_dfm:
            continue
        try:
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        except Exception:
            pass
    trim_trailing_rows(ws, last_data_row)


def fill_details(ws, details: list[dict[str, Any]]) -> None:
    """4.상세코드 시트에 상세 데이터 작성."""
    START_ROW = 4
    MAX_COL = 9
    template_last_row = ws.max_row  # 템플릿 데이터 영역의 마지막 행 (스타일 보유 한계)
    clear_data_rows(ws, START_ROW, MAX_COL)

    for idx, d in enumerate(details):
        r = START_ROW + idx
        biz_seq = d.get("biz_seq")
        # 템플릿이 문자열로 1, 2 ... 를 넣어둠 → 일관성 유지를 위해 문자열로
        ws.cell(row=r, column=1, value=str(biz_seq) if biz_seq is not None else None)
        ws.cell(row=r, column=2, value=d.get("comm_h_cd"))
        ws.cell(row=r, column=3, value=d.get("comm_d_cd"))
        ws.cell(row=r, column=4, value=d.get("comm_d_nm"))
        ws.cell(row=r, column=5, value=d.get("ref_h_cd") or None)
        ws.cell(row=r, column=6, value=d.get("ref_d_cd") or None)
        # G=disp_no (정수)
        disp_no = d.get("disp_no")
        try:
            ws.cell(row=r, column=7, value=int(disp_no) if disp_no is not None else None)
        except (TypeError, ValueError):
            ws.cell(row=r, column=7, value=disp_no)
        ws.cell(row=r, column=8, value=d.get("disp_yn") or "Y")
        ws.cell(row=r, column=9, value=d.get("use_yn") or "Y")
        # 템플릿 행 수를 초과한 새 행은 첫 데이터 행 스타일/행 높이를 복사한다.
        if r > template_last_row:
            copy_row_style(ws, START_ROW, r, MAX_COL)

    last_data_row = START_ROW + len(details) - 1 if details else START_ROW - 1
    trim_trailing_rows(ws, last_data_row)


def fill_sql_sheet(ws, count: int, formula_fn, b_text: str, c_text: str) -> None:
    """SQL 시트의 행을 데이터 건수에 맞춰 재구성한다."""
    template_last_row = ws.max_row  # 템플릿 데이터 영역의 마지막 행 (스타일 보유 한계)
    SQL_MAX_COL = 3  # A, B, C
    # 모든 기존 행을 비운다 (value만 None으로)
    if template_last_row >= 1:
        for r in range(1, template_last_row + 1):
            for c in range(1, SQL_MAX_COL + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    cell.value = None
    # 새로 채움 — 템플릿 한계를 넘어선 행은 1행 스타일을 복사한다.
    for n in range(1, count + 1):
        ws.cell(row=n, column=1, value=formula_fn(n))
        ws.cell(row=n, column=2, value=b_text)
        ws.cell(row=n, column=3, value=c_text)
        if n > template_last_row:
            copy_row_style(ws, 1, n, SQL_MAX_COL)
    # count 이후의 빈 행을 물리적으로 제거
    trim_trailing_rows(ws, max(count, 1))


def main() -> int:
    if not IN_FILE.exists():
        print(f"[ERROR] {IN_FILE} 가 없습니다. 2단계(공통코드 추출)를 먼저 수행하세요.", file=sys.stderr)
        return 2
    if not TEMPLATE.exists():
        print(f"[ERROR] 템플릿이 없습니다: {TEMPLATE}", file=sys.stderr)
        return 2

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        print("[ERROR] openpyxl 미설치: pip install --user openpyxl", file=sys.stderr)
        return 3

    payload = json.loads(IN_FILE.read_text(encoding="utf-8"))
    groups = payload.get("groups", [])
    details = payload.get("details", [])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    yymmdd = datetime.now().strftime("%y%m%d")
    out_path = OUT_DIR / f"PI_113-공통코드정의서_{yymmdd}.xlsx"

    shutil.copyfile(TEMPLATE, out_path)
    wb = load_workbook(out_path)

    # 메타정보: 3.코드그룹 작성일자(E4) — 시간 제외, yyyy-mm-dd 형식으로 통일
    if "3.코드그룹" in wb.sheetnames:
        ws_meta = wb["3.코드그룹"]
        try:
            cell = ws_meta.cell(row=4, column=5, value=datetime.now().date())
            cell.number_format = "yyyy-mm-dd"
        except Exception:
            pass

    # 데이터 시트 채우기
    if "3.코드그룹" in wb.sheetnames:
        fill_groups(wb["3.코드그룹"], groups)
    if "4.상세코드" in wb.sheetnames:
        fill_details(wb["4.상세코드"], details)

    # SQL 시트
    if "그룹SQL" in wb.sheetnames:
        fill_sql_sheet(wb["그룹SQL"], len(groups), make_group_formula, GROUP_B_TEXT, GROUP_C_TEXT)
    if "상세SQL" in wb.sheetnames:
        fill_sql_sheet(wb["상세SQL"], len(details), make_detail_formula, DETAIL_B_TEXT, DETAIL_C_TEXT)

    wb.save(out_path)

    used = sum(1 for g in groups if (g.get("use_yn") or "Y").upper() == "Y")
    unused = len(groups) - used
    print(f"[OK] 생성 완료: {out_path}")
    print(f"  - 코드그룹(SM_COMM_H): {len(groups)}건 (사용 {used} / 미사용 {unused})")
    print(f"  - 상세코드(SM_COMM_D): {len(details)}건")
    return 0


if __name__ == "__main__":
    sys.exit(main())
