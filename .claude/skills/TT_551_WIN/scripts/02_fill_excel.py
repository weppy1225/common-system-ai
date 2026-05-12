#!/usr/bin/env python3
"""
TT_551_WIN 3·4단계 - 템플릿 복사 + sentinel 치환 + 표 반복 행 채우기.

사용법:
    python 02_fill_excel.py <template_xlsx> <output_xlsx> <fill_data_json>

동작:
  1) template_xlsx 를 output_xlsx 로 복사
  2) 모든 시트를 순회하며 {{key}} 스칼라 sentinel 을 fill_data.scalars 의 값으로 치환
  3) {{rows:groups}}, {{rows:tables}} 마커가 있는 행을 첫 데이터 행으로 보고
     fill_data.rows[xxx] 의 dict 리스트를 한 행씩 채워 넣음 (행 자동 삽입, 서식 보존)

요구사항: openpyxl
"""

import json
import shutil
import sys
from pathlib import Path
from copy import copy

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl 가 설치되어 있지 않습니다.", file=sys.stderr)
    print("       pip install --user openpyxl", file=sys.stderr)
    sys.exit(2)


SENTINEL_OPEN = "{{"
SENTINEL_CLOSE = "}}"
ROWS_PREFIX = "rows:"


def replace_scalars(ws, scalars: dict) -> int:
    """모든 셀의 텍스트 안의 {{key}} 패턴을 scalars[key] 로 치환. 치환 건수 반환."""
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            v = cell.value
            if SENTINEL_OPEN not in v:
                continue
            for key, val in scalars.items():
                token = SENTINEL_OPEN + key + SENTINEL_CLOSE
                if token in v:
                    v = v.replace(token, str(val) if val is not None else "")
                    count += 1
            cell.value = v
    return count


def find_row_markers(ws):
    """{{rows:xxx}} 마커가 들어 있는 셀들을 (cell, key) 튜플 리스트로 반환."""
    found = []
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            stripped = cell.value.strip()
            if (stripped.startswith(SENTINEL_OPEN + ROWS_PREFIX)
                    and stripped.endswith(SENTINEL_CLOSE)):
                key = stripped[len(SENTINEL_OPEN) + len(ROWS_PREFIX):-len(SENTINEL_CLOSE)]
                found.append((cell, key))
    return found


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def fill_table_rows(ws, marker_cell, data_rows):
    """marker_cell 이 있는 행을 첫 데이터 행으로 보고, data_rows 를 한 행씩 채운다.

    - 컬럼 순서: data_rows[0] 의 dict 키 순서대로 marker_cell 부터 우측으로.
    - 행이 부족하면 ws.insert_rows() 로 늘림. 기존 서식 복제.
    """
    if not data_rows:
        # 마커만 비우고 끝
        marker_cell.value = None
        return

    start_row = marker_cell.row
    start_col = marker_cell.column
    keys = list(data_rows[0].keys())
    n_rows = len(data_rows)

    # 마커 행 기준 셀들의 서식을 미리 저장 (n_rows-1 개 행 추가 후 복제용)
    template_cells = []
    for i, key in enumerate(keys):
        c = ws.cell(row=start_row, column=start_col + i)
        template_cells.append(c)

    # 1행 초과면 그만큼 아래로 행 삽입 (start_row + 1 위치에 n_rows-1 개)
    if n_rows > 1:
        ws.insert_rows(start_row + 1, amount=n_rows - 1)

    # 데이터 채우기
    for r_idx, data_row in enumerate(data_rows):
        target_row = start_row + r_idx
        for c_idx, key in enumerate(keys):
            cell = ws.cell(row=target_row, column=start_col + c_idx)
            # 첫 행은 마커 셀 위치이므로 서식이 이미 있음
            # 신규 행에는 template_cells 서식 복제
            if r_idx > 0:
                copy_cell_style(template_cells[c_idx], cell)
            val = data_row.get(key)
            cell.value = val if val is not None else ""


def main():
    if len(sys.argv) < 4:
        print("usage: 02_fill_excel.py <template_xlsx> <output_xlsx> <fill_data_json>", file=sys.stderr)
        sys.exit(2)

    template_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    fill_data_path = Path(sys.argv[3])

    if not template_path.exists():
        print(f"ERROR: 템플릿이 없습니다: {template_path}", file=sys.stderr)
        sys.exit(2)

    fill_data = json.loads(fill_data_path.read_text(encoding="utf-8"))
    scalars = fill_data.get("scalars", {})
    rows_data = fill_data.get("rows", {})

    # 1) 템플릿 → 출력 파일 복사
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    # 2) 워크북 열기
    wb = openpyxl.load_workbook(output_path)

    total_scalar_replacements = 0
    total_row_fills = 0
    found_markers = []

    # 3) 표 마커 먼저 찾기 — 행 삽입이 일어나면 스칼라 sentinel 위치가 밀려서 그 다음에 치환
    for ws in wb.worksheets:
        markers = find_row_markers(ws)
        for cell, key in markers:
            data = rows_data.get(key)
            if data is None:
                print(f"[TT_551_WIN] ⚠ rows:{key} 데이터가 fill_data 에 없습니다. 마커 셀 비우고 진행.", file=sys.stderr)
                cell.value = None
                continue
            n = len(data)
            print(f"[TT_551_WIN] {ws.title}!{cell.coordinate}: rows:{key} → {n}행 채움")
            fill_table_rows(ws, cell, data)
            total_row_fills += n
            found_markers.append((ws.title, key, n))

    # 4) 스칼라 sentinel 치환 (행 삽입 후)
    for ws in wb.worksheets:
        total_scalar_replacements += replace_scalars(ws, scalars)

    wb.save(output_path)

    size_kb = round(output_path.stat().st_size / 1024, 1)
    print(f"[TT_551_WIN] 엑셀 채우기 완료: {output_path}")
    print(f"  스칼라 치환:  {total_scalar_replacements}건")
    print(f"  표 행 채움:   {total_row_fills}건  ({len(found_markers)} 마커)")
    print(f"  파일 크기:    {size_kb} KB")


if __name__ == "__main__":
    main()
