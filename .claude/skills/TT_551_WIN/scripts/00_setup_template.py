#!/usr/bin/env python3
"""TT_551_WIN 사전 준비 — 템플릿에 sentinel placeholder를 자동 삽입.

input/TT.551/TT_551_DB이관계획서_20260511.xlsx 기반 구조 가정.

사용법:
    python 00_setup_template.py <template_xlsx>
"""

import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


THIN = Side(style="thin", color="C0C0C0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_cell(ws, coord, value, *, bold=False, fill_color=None, align="left"):
    cell = ws[coord]
    cell.value = value
    cell.font = Font(name="맑은 고딕", size=10, bold=bold)
    cell.alignment = Alignment(
        horizontal=align,
        vertical="center",
        wrap_text=True,
    )
    cell.border = BOX
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)


def setup_cover(wb):
    """표지 시트 — M9 아래로 고객사·작성자·작성일 sentinel 추가."""
    ws = wb["표지"]
    # M11~N15 영역에 라벨 + sentinel 추가
    set_cell(ws, "M11", "고객사", bold=True, fill_color="F3F4F6", align="center")
    set_cell(ws, "M12", "{{customer}}", align="center")
    set_cell(ws, "M14", "작성자", bold=True, fill_color="F3F4F6", align="center")
    set_cell(ws, "M15", "{{author}}", align="center")
    set_cell(ws, "M17", "작성일", bold=True, fill_color="F3F4F6", align="center")
    set_cell(ws, "M18", "{{generated_at}}", align="center")


def setup_overview(wb):
    """이관개요 시트 — 기존 행에 sentinel을 박고 추가 행을 덧붙임."""
    ws = wb["이관개요"]

    # 기존 셀의 값을 sentinel로 교체
    ws["B5"].value = "{{plan_date}}"
    ws["B5"].number_format = "@"  # 텍스트 포맷
    ws["B7"].value = "{{author}}"
    ws["B8"].value = "{{source_host}}"
    ws["B10"].value = "{{source_db}}"

    # 추가 행 (A19~B27)
    template_th = ws["A2"]  # 기존 헤더 셀 (서식 복제용)
    template_td = ws["B2"]

    additions = [
        (19, "원본 schema", "{{source_schema}}"),
        (20, "PG 버전", "{{source_pg_version}}"),
        (21, "그룹 수", "{{group_count}}"),
        (22, "테이블 수", "{{table_count}}"),
        (23, "전체 row 수", "{{row_count_total}}"),
        (24, "경고 (마커 누락)", "{{warnings}}"),
        (25, "사전 조건", "{{precondition}}"),
        (26, "롤백 계획", "{{rollback}}"),
    ]

    for row_idx, label, value in additions:
        a_cell = ws.cell(row=row_idx, column=1, value=label)
        b_cell = ws.cell(row=row_idx, column=2, value=value)
        # 서식 복제 (이관개요 기존 데이터행 스타일)
        if template_th.has_style:
            a_cell.font = copy(template_th.font)
            a_cell.fill = copy(template_th.fill)
            a_cell.border = copy(template_th.border)
            a_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
        if template_td.has_style:
            b_cell.font = copy(template_td.font)
            b_cell.fill = copy(template_td.fill)
            b_cell.border = copy(template_td.border)
            b_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

    # 사전조건/롤백/warnings는 multi-line 가능성 → 행 높이 확장
    for r in (24, 25, 26):
        ws.row_dimensions[r].height = 48


def setup_tables_list(wb):
    """이관대상목록 시트 — 헤더를 표준 8컬럼으로 재정의 + A3에 {{rows:tables}} 마커."""
    ws = wb["이관대상목록"]

    # 기존 데이터 행 삭제 (A3:H14 → 12행)
    ws.delete_rows(3, ws.max_row - 2)

    # 새 헤더 (A2:H2)
    new_headers = [
        ("A2", "순번"),
        ("B2", "분류"),
        ("C2", "테이블명"),
        ("D2", "테이블설명"),
        ("E2", "row수"),
        ("F2", "FK부모"),
        ("G2", "SQL파일"),
        ("H2", "적용순서"),
    ]
    for coord, text in new_headers:
        cell = ws[coord]
        cell.value = text
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="304A6E")
        cell.fill = PatternFill("solid", fgColor="E0E7F0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX

    # 컬럼 너비 조정
    widths = {"A": 6, "B": 12, "C": 18, "D": 22, "E": 8, "F": 16, "G": 24, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 마커 행 (A3에 {{rows:tables}}, B3~H3은 빈 칸이지만 서식 적용)
    align_map = {
        "A": "center",
        "B": "center",
        "C": "left",
        "D": "left",
        "E": "right",
        "F": "left",
        "G": "left",
        "H": "center",
    }
    for col_letter, align in align_map.items():
        cell = ws[f"{col_letter}3"]
        cell.font = Font(name="맑은 고딕", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = BOX
    ws["A3"].value = "{{rows:tables}}"


def setup_groups_summary(wb):
    """새 시트 '이관그룹요약' 추가 + {{rows:groups}} 마커."""
    sheet_name = "이관그룹요약"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name, index=wb.sheetnames.index("이관대상목록"))

    # 제목
    ws["A1"] = "■ 이관 그룹 요약"
    ws["A1"].font = Font(name="맑은 고딕", size=12, bold=True, color="304A6E")
    ws.merge_cells("A1:G1")

    # 헤더 (A2:G2) - 7컬럼: no, group_key, group_desc, table_count, row_sum, sql_file, file_size_kb
    headers = [
        ("A2", "순번"),
        ("B2", "그룹키"),
        ("C2", "그룹명"),
        ("D2", "테이블 수"),
        ("E2", "row 합계"),
        ("F2", "SQL 파일"),
        ("G2", "파일 크기(KB)"),
    ]
    for coord, text in headers:
        cell = ws[coord]
        cell.value = text
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="304A6E")
        cell.fill = PatternFill("solid", fgColor="E0E7F0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX

    widths = {"A": 6, "B": 18, "C": 14, "D": 10, "E": 10, "F": 22, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    align_map = {
        "A": "center",
        "B": "left",
        "C": "left",
        "D": "right",
        "E": "right",
        "F": "left",
        "G": "right",
    }
    for col_letter, align in align_map.items():
        cell = ws[f"{col_letter}3"]
        cell.font = Font(name="맑은 고딕", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = BOX
    ws["A3"].value = "{{rows:groups}}"


def main():
    if len(sys.argv) < 2:
        print("usage: 00_setup_template.py <template_xlsx>", file=sys.stderr)
        sys.exit(2)

    template_path = Path(sys.argv[1])
    if not template_path.exists():
        print(f"ERROR: 템플릿이 없습니다: {template_path}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(template_path)

    setup_cover(wb)
    setup_overview(wb)
    setup_groups_summary(wb)
    setup_tables_list(wb)

    wb.save(template_path)
    print(f"[TT_551_WIN] 템플릿 sentinel 셋업 완료: {template_path}")

    # 검증
    wb2 = openpyxl.load_workbook(template_path)
    sentinels = []
    for ws in wb2.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    sentinels.append((ws.title, cell.coordinate, cell.value))
    print(f"  박힌 sentinel: {len(sentinels)}개")
    for s in sentinels:
        print(f"    {s[0]}!{s[1]}: {s[2]}")


if __name__ == "__main__":
    main()
